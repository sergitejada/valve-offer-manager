import openpyxl
from openpyxl.utils import column_index_from_string

template_path = "offer-template.xlsx"
output_path = "items_output_2.xlsx"  

wb = openpyxl.load_workbook(template_path)
ws = wb.worksheets[1] 

# JSON de entrada
input_json = [
    {
        "item": 1,
        "valve type": "ball valve",
        "end_user": "User1",
        "9COM": "Code1",
        "size": "Small",
        "class": "A",
        "operation": "Open",
        "body_construction": "Steel",
        "ball_construction": "Plastic",
        "bore": "10",
        "ends": "Welded",
        "standard": "ISO",
        "tag": "Tag1",
        "model": "Model1",
        "body": "Body1",
        "ball": "Ball1",
        "stem": "Stem1",
        "seat_housing": "Housing1",
        "seat": "Seat1",
        "seals": "Seal1",
        "injectors": "Injector1",
        "drain_vent": "Vent1",
        "painting": "Paint1",
        "temperature": "20C",
        "service": "Service1",
        "available": "Yes"
    },
    {
        "item": 2,
        "valve type": "ball valve",
        "end_user": "User2",
        "9COM": "Code2",
        "size": "Medium",
        "class": "B",
        "operation": "Close",
        "body_construction": "Aluminum",
        "ball_construction": "Metal",
        "bore": "20",
        "ends": "Flanged",
        "standard": "ANSI",
        "tag": "Tag2",
        "model": "Model2",
        "body": "Body2",
        "ball": "Ball2",
        "stem": "Stem2",
        "seat_housing": "Housing2",
        "seat": "Seat2",
        "seals": "Seal2",
        "injectors": "Injector2",
        "drain_vent": "Vent2",
        "painting": "Paint2",
        "temperature": "25C",
        "service": "Service2",
        "available": "No"
    }    
]

items_to_excel = {
    "valve type": column_index_from_string("E"),
    "end_user": column_index_from_string("D"),
    "9COM": column_index_from_string("J"),
    "size": column_index_from_string("K"),
    "class": column_index_from_string("L"),
    "operation": column_index_from_string("F"),
    "body_construction": column_index_from_string("M"),
    "ball_construction": column_index_from_string("N"),
    "bore": column_index_from_string("O"),
    "ends": column_index_from_string("P"),
    "standard": column_index_from_string("Q"),
    "tag": column_index_from_string("G"),
    "model": column_index_from_string("R"),
    "body": column_index_from_string("S"),
    "ball": column_index_from_string("T"),
    "stem": column_index_from_string("U"),
    "seat_housing": column_index_from_string("V"),
    "seat": column_index_from_string("W"),
    "seals": column_index_from_string("X"),
    "injectors": column_index_from_string("Y"),
    "drain_vent": column_index_from_string("Z"),
    "painting": column_index_from_string("AA"),
    "temperature": column_index_from_string("AB"),
    "service": column_index_from_string("AC"),
    "available": column_index_from_string("AS")
}

starting_row = 9  
for index, item in enumerate(input_json, start=0):

    ws.cell(row=starting_row + index, column=2, value=item["item"])
    

    for attribute, col_index in items_to_excel.items():

        value = item.get(attribute, "N/A")
        

        ws.cell(row=starting_row + index, column=col_index, value=value)


wb.save(output_path)

print(f"Arxiu Excel Creat: {output_path}")
